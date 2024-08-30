import streamlit as st
import pandas as pd
import csv
import io
from scraper import scrape_clinic_site
from preprocessor import preprocess_data
from seo_optimizer import generate_seo_proposals
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# パスワード認証機能を追加
def check_password():
    """Returns `True` if the user had the correct password."""

    def password_entered():
        """Checks whether a password entered by the user is correct."""
        if st.session_state["password"] == st.secrets["password"]:
            st.session_state["password_correct"] = True
            del st.session_state["password"]  # don't store password
        else:
            st.session_state["password_correct"] = False

    if "password_correct" not in st.session_state:
        # First run, show input for password.
        st.text_input(
            "パスワードを入力してください", type="password", on_change=password_entered, key="password"
        )
        return False
    elif not st.session_state["password_correct"]:
        # Password not correct, show input + error.
        st.text_input(
            "パスワードを入力してください", type="password", on_change=password_entered, key="password"
        )
        st.error("😕 パスワードが正しくありません")
        return False
    else:
        # Password correct.
        return True

def main():
    st.title("クリニックSEO最適化支援ツール")
    st.markdown("✨更新情報  \n▼ver1.0.0  \nツール作成しました。")
    st.markdown("<font color='red' style='font-size: 13px;line-height:1.5;'>URLにblogが入っている場合は検索対象になりません（膨大な量になる可能性があるため）。<br>そのため、もしも診療案内としてblogを使用されている医院がありましたら、<br>お手数ですが、手動にてご確認をお願いいたします。</font>", unsafe_allow_html=True)
    st.markdown("※必ず、人の確認をいれてください。  \n※GPT4o-miniを使用しています。個人情報の入力はご遠慮ください。  \n※外部LLMを使用しているため料金が発生します。計画的にお使いください。")
    
    # セッションステートの初期化
    if 'seo_proposals' not in st.session_state:
        st.session_state.seo_proposals = None
    if 'analysis_complete' not in st.session_state:
        st.session_state.analysis_complete = False
    if 'clinic_name' not in st.session_state:
        st.session_state.clinic_name = ""

    # ユーザー入力
    clinic_url = st.text_input("クリニックサイトのURLを入力してください")
    seo_goal = st.text_area("SEO最適化の目標を入力してください")

    if st.button("分析開始"):
        if clinic_url and seo_goal:
            try:
                # 進捗バーの初期化
                progress_bar = st.progress(0)
                status_text = st.empty()

                # スクレイピング
                scraped_data = scrape_clinic_site(clinic_url, lambda p, m: update_progress(progress_bar, status_text, p, m))
                
                # クリニック名の抽出
                st.session_state.clinic_name = extract_clinic_name(scraped_data)
                
                # データ前処理
                update_progress(progress_bar, status_text, 0.7, "データ前処理中...")
                processed_data = preprocess_data(scraped_data)
                
                # SEO最適化提案生成
                update_progress(progress_bar, status_text, 0.8, "SEO最適化提案生成中(GPT-4o-mini)...")
                st.session_state.seo_proposals = generate_seo_proposals(processed_data, seo_goal)
                
                # 分析完了フラグを設定
                st.session_state.analysis_complete = True
                
                # 結果表示
                update_progress(progress_bar, status_text, 1.0, "完了")

            except Exception as e:
                st.error(f"エラーが発生しました: {str(e)}")
        else:
            st.error("URLとSEO目標を入力してください。")

    # 結果の表示（セッションステートを使用）
    if st.session_state.analysis_complete:
        display_results(st.session_state.seo_proposals)

        # Excelファイルダウンロードボタンの追加
        excel = convert_to_excel(st.session_state.seo_proposals)
        file_name = f"{st.session_state.clinic_name}_SEO最適化案.xlsx"
        st.download_button(
            label="結果をExcelでダウンロード",
            data=excel,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

def update_progress(progress_bar, status_text, progress, message):
    progress_bar.progress(progress)
    status_text.text(message)

def display_results(seo_proposals):
    st.subheader("SEO最適化提案")
    for page, proposals in seo_proposals.items():
        with st.expander(f"ページ: {page}"):
            st.markdown("### 現在の情報:", unsafe_allow_html=True)
            st.write(f"タイトル: {proposals['current_title']} (文字数: {len(proposals['current_title'])})")
            st.write(f"ディスクリプション: {proposals['current_description']} (文字数: {len(proposals['current_description'])})")
            
            st.markdown("### 最適化提案:", unsafe_allow_html=True)
            for i, title in enumerate(proposals['proposed_titles'], 1):
                st.write(f"タイトル案 {i}: {title} (文字数: {len(title)})")
            for i, desc in enumerate(proposals['proposed_descriptions'], 1):
                st.write(f"ディスクリプション案 {i}: {desc} (文字数: {len(desc)})")

def convert_to_excel(seo_proposals):
    wb = Workbook()
    ws = wb.active
    row_num = 1

    for page, proposals in seo_proposals.items():
        ws.cell(row=row_num, column=1, value='ページURL')
        ws.cell(row=row_num, column=2, value=page)
        ws.cell(row=row_num, column=3, value='文字数')
        row_num += 1

        ws.cell(row=row_num, column=1, value='現在のタイトル')
        ws.cell(row=row_num, column=2, value=proposals.get('current_title', ''))
        ws.cell(row=row_num, column=3, value=f'=LEN(B{row_num})')
        row_num += 1
        
        for i, title in enumerate(proposals.get('proposed_titles', []), 1):
            ws.cell(row=row_num, column=1, value=f'タイトル案{i}')
            ws.cell(row=row_num, column=2, value=title)
            ws.cell(row=row_num, column=3, value=f'=LEN(B{row_num})')
            row_num += 1
        
        ws.cell(row=row_num, column=1, value='現在のディスクリプション')
        ws.cell(row=row_num, column=2, value=proposals.get('current_description', ''))
        ws.cell(row=row_num, column=3, value=f'=LEN(B{row_num})')
        row_num += 1
        
        for i, desc in enumerate(proposals.get('proposed_descriptions', []), 1):
            ws.cell(row=row_num, column=1, value=f'ディスクリプション案{i}')
            ws.cell(row=row_num, column=2, value=desc)
            ws.cell(row=row_num, column=3, value=f'=LEN(B{row_num})')
            row_num += 1
        
        row_num += 1  # 空行を挿入

    # 列幅の自動調整
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file.getvalue()

def extract_clinic_name(scraped_data):
    # トップページのタイトルからクリニック名を抽出
    for url, data in scraped_data.items():
        if url.endswith('/') or url.endswith('/index.html'):
            title = data['title']
            # クリニック名を抽出する正規表現パターン
            match = re.search(r'(.+?)(クリニック|病院|医院)', title)
            if match:
                return match.group(0)
    return "クリニック"  # デフォルト名

if __name__ == "__main__":
    if check_password():
        main()